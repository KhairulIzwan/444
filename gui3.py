#!/usr/bin/env python

from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

# read  from excel file
wb = load_workbook('/Users/khairulizwan/Scripts/444/Original Data/444_1.xlsx')
sheet_1 = wb.get_sheet_by_name('case1')

plt.figure(figsize=(6, 4), facecolor='Grey')
G = gridspec.GridSpec(6, 2)
axes_1 = plt.subplot(G[0, :])


x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=4).value

#print x
#print y

# create the plot
plt.xlabel('time')
plt.ylabel('HR')
plt.plot(x, y, color='cyan', label='HR')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)
#plt.title('Reading values from an Excel file'

axes_1 = plt.subplot(G[1, :])

x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=6).value

#print a
#print b

# create the plot
plt.xlabel('time')
plt.ylabel('Pulse')
plt.plot(x, y, color='red', label='Pulse')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)

axes_1 = plt.subplot(G[2, :])

x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=7).value

#print x
#print y

# create the plot
plt.xlabel('time')
plt.ylabel('SpO2')
plt.plot(x, y, color='magenta', label='SpO2')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)

axes_1 = plt.subplot(G[3, :])

x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=8).value

#print x
#print y

# create the plot
plt.xlabel('time')
plt.ylabel('Perf')
plt.plot(x, y, color='blue', label='Perf')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)


axes_1 = plt.subplot(G[4, :])

x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=9).value

#print x
#print y

# create the plot
plt.xlabel('time')
plt.ylabel('etCO2')
plt.plot(x, y, color='yellow', label='etCO2')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)


axes_1 = plt.subplot(G[5, :])

x = np.zeros(sheet_1.max_row)
y = np.zeros(len(x))

for i in range(1, sheet_1.max_row):
    x[i] = sheet_1.cell(row=i + 1, column=2).value
    y[i] = sheet_1.cell(row=i + 1, column=10).value

#print x
#print y

# create the plot
plt.xlabel('time')
plt.ylabel('imCO2')
plt.plot(x, y, color='green', label='imCO2')
plt.legend(loc='upper right', fontsize='small')
plt.grid(True)

plt.xlim(0, 60000)
plt.ylim(0, 100)
plt.show()
